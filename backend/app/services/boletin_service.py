from datetime import datetime, date
from io import BytesIO
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel as excel_serial_to_datetime
from sqlalchemy.orm import Session

from app.models.anums import ModuloEnum, ImpactoEnum
from app.models.boletines import Boletin
from app.repositories.boletin_repository import BoletinRepository
from app.schemas.boletin import BoletinCreate


class BoletinService:

    REQUIRED_COLUMNS = [
        "tipo de documento",
        "consecutivo",
        "fecha",
        "modulo",
        "opcion",
        "impacto",
        "categoria",
        "con documentacion",
        "asunto",
        "clase de documento",
        "advertencia",
        "instructivos descripcion",
    ]

    def __init__(self):
        self.repository = BoletinRepository()

    def listar(self, db: Session, mes: int | None = None, anio: int | None = None):
        return self.repository.get_all(db, mes=mes, anio=anio)

    def listar_periodos(self, db: Session):
        rows = self.repository.list_periodos(db)
        return [{"mes": int(row.mes), "anio": int(row.anio)} for row in rows]

    @staticmethod
    def _format_date(value: datetime | None) -> str:
        if value is None:
            return ""
        return value.strftime("%Y-%m-%d")

    @staticmethod
    def _matches_filter(value, query: str | None) -> bool:
        expected = (query or "").strip().lower()
        if not expected:
            return True
        candidate = "" if value is None else str(value).strip().lower()
        return expected in candidate

    def _filter_boletines(
        self,
        boletines: list[Boletin],
        *,
        consecutivo: str | None = None,
        modulo: str | None = None,
        fecha: str | None = None,
        opcion: str | None = None,
        impacto: str | None = None,
        categoria: str | None = None,
        clase_documento: str | None = None,
        asunto: str | None = None,
    ) -> list[Boletin]:
        filtered: list[Boletin] = []
        for boletin in boletines:
            if not self._matches_filter(boletin.consecutivo, consecutivo):
                continue
            if not self._matches_filter(boletin.modulo, modulo):
                continue
            if not self._matches_filter(self._format_date(boletin.fecha), fecha):
                continue
            if not self._matches_filter(boletin.opcion, opcion):
                continue
            if not self._matches_filter(boletin.impacto, impacto):
                continue
            if not self._matches_filter(boletin.categoria, categoria):
                continue
            if not self._matches_filter(boletin.clase_documento, clase_documento):
                continue
            if not self._matches_filter(boletin.asunto, asunto):
                continue
            filtered.append(boletin)
        return filtered

    def exportar_excel_filtrado(
        self,
        db: Session,
        *,
        mes: int,
        anio: int,
        consecutivo: str | None = None,
        modulo: str | None = None,
        fecha: str | None = None,
        opcion: str | None = None,
        impacto: str | None = None,
        categoria: str | None = None,
        clase_documento: str | None = None,
        asunto: str | None = None,
    ) -> tuple[BytesIO, str]:
        boletines = self.repository.get_all(db, mes=mes, anio=anio)
        filtered = self._filter_boletines(
            boletines,
            consecutivo=consecutivo,
            modulo=modulo,
            fecha=fecha,
            opcion=opcion,
            impacto=impacto,
            categoria=categoria,
            clase_documento=clase_documento,
            asunto=asunto,
        )
        if not filtered:
            raise ValueError("No hay boletines para exportar con los filtros seleccionados.")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"{anio}-{mes:02d}"

        header_fill = PatternFill(start_color="0778AC", end_color="0778AC", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        headers = [
            "Tipo de documento",
            "Consecutivo",
            "Fecha",
            "Modulo",
            "Opcion",
            "Impacto",
            "Categoria",
            "Con documentacion",
            "Asunto",
            "Clase de documento",
            "Advertencia",
            "Instructivos descripcion",
            "Mes",
            "Año",
        ]
        sheet.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for boletin in filtered:
            sheet.append([
                boletin.tipo_documento or "",
                boletin.consecutivo or "",
                self._format_date(boletin.fecha),
                boletin.modulo.value if hasattr(boletin.modulo, "value") else (boletin.modulo or ""),
                boletin.opcion or "",
                boletin.impacto.value if hasattr(boletin.impacto, "value") else (boletin.impacto or ""),
                boletin.categoria or "",
                "Sí" if boletin.con_documentacion else "No",
                boletin.asunto or "",
                boletin.clase_documento or "",
                boletin.advertencia or "",
                boletin.instructivo_descripcion or "",
                boletin.mes,
                boletin.anio,
            ])

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            column_letter = get_column_letter(column[0].column)
            sheet.column_dimensions[column_letter].width = min(max(max_len + 4, 14), 50)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"boletines_filtrados_{anio}_{mes:02d}.xlsx"
        return output, filename

    @staticmethod
    def _canon(value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value)
        without_accents = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        no_symbols = re.sub(r"[^a-zA-Z0-9\s]", " ", without_accents)
        return re.sub(r"\s+", " ", no_symbols).strip().lower()

    @staticmethod
    def _to_bool(value) -> bool:
        if isinstance(value, bool):
            return value
        if value is None:
            return False
        text = str(value).strip().lower()
        return text in {"si", "sí", "s", "1", "true", "x", "yes"}

    @staticmethod
    def _to_int(value):
        if value is None or str(value).strip() == "":
            return None
        try:
            return int(float(str(value).strip()))
        except Exception:
            # Keep import resilient: when consecutivo is not numeric, persist null.
            return None

    @staticmethod
    def _to_date(value):
        if value is None or str(value).strip() == "":
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        if isinstance(value, (int, float)):
            try:
                excel_dt = excel_serial_to_datetime(value)
                if isinstance(excel_dt, datetime):
                    return excel_dt
                if isinstance(excel_dt, date):
                    return datetime.combine(excel_dt, datetime.min.time())
            except Exception:
                return None

        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def _to_modulo(value):
        if value is None or str(value).strip() == "":
            return None
        candidate = BoletinService._canon(str(value)).replace(" ", "_").upper()
        valid = {item.value for item in ModuloEnum}
        if candidate not in valid:
            return "OTROS"
        return candidate

    @staticmethod
    def _to_impacto(value):
        if value is None or str(value).strip() == "":
            return None
        candidate = BoletinService._canon(str(value))
        aliases = {
            "high": "alto",
            "medium": "medio",
            "low": "bajo",
        }
        candidate = aliases.get(candidate, candidate)
        valid = {item.value for item in ImpactoEnum}
        if candidate not in valid:
            return None
        return candidate

    def importar_excel(self, db: Session, excel_bytes: bytes, mes: int, anio: int):
        if mes < 1 or mes > 12:
            raise ValueError("El mes debe estar entre 1 y 12.")
        if anio < 2000 or anio > 2100:
            raise ValueError("El año debe estar entre 2000 y 2100.")

        workbook = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
        sheet = workbook.active

        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            raise ValueError("El archivo Excel está vacío.")

        header_map = {}
        for index, cell in enumerate(header_row):
            if cell is None:
                continue
            key = self._canon(str(cell))
            header_map[key] = index

        missing = [column for column in self.REQUIRED_COLUMNS if column not in header_map]
        if missing:
            raise ValueError(f"Faltan columnas requeridas en Excel: {', '.join(missing)}")

        boletines = []
        skipped_rows = 0
        for row_number, row in enumerate(rows_iter, start=2):
            if row is None:
                continue

            tipo_documento = row[header_map["tipo de documento"]]
            consecutivo = row[header_map["consecutivo"]]
            fecha = row[header_map["fecha"]]
            modulo = row[header_map["modulo"]]
            opcion = row[header_map["opcion"]]
            impacto = row[header_map["impacto"]]
            categoria = row[header_map["categoria"]]
            con_documentacion = row[header_map["con documentacion"]]
            asunto = row[header_map["asunto"]]
            clase_documento = row[header_map["clase de documento"]]
            advertencia = row[header_map["advertencia"]]
            instructivo_descripcion = row[header_map["instructivos descripcion"]]

            values = [
                tipo_documento,
                consecutivo,
                fecha,
                modulo,
                opcion,
                impacto,
                categoria,
                con_documentacion,
                asunto,
                clase_documento,
                advertencia,
                instructivo_descripcion,
            ]
            if all(value is None or str(value).strip() == "" for value in values):
                continue

            try:
                boletines.append(
                    Boletin(
                        tipo_documento=str(tipo_documento).strip() if tipo_documento is not None else None,
                        consecutivo=self._to_int(consecutivo),
                        fecha=self._to_date(fecha),
                        modulo=self._to_modulo(modulo),
                        opcion=str(opcion).strip() if opcion is not None else None,
                        impacto=self._to_impacto(impacto),
                        categoria=str(categoria).strip() if categoria is not None else None,
                        con_documentacion=self._to_bool(con_documentacion),
                        asunto=str(asunto).strip() if asunto is not None else None,
                        clase_documento=str(clase_documento).strip() if clase_documento is not None else None,
                        advertencia=str(advertencia).strip() if advertencia is not None else None,
                        instructivo_descripcion=str(instructivo_descripcion).strip() if instructivo_descripcion is not None else None,
                        mes=mes,
                        anio=anio,
                        archivo=None,
                        fecha_registro=datetime.now(),
                    )
                )
            except Exception:
                skipped_rows += 1
                continue

        if not boletines:
            raise ValueError("No se encontraron filas con datos para importar.")

        self.repository.delete_by_period(db, mes=mes, anio=anio)
        self.repository.create_many(db, boletines)
        return {
            "mes": mes,
            "anio": anio,
            "inserted_rows": len(boletines),
            "skipped_rows": skipped_rows,
        }

    def crear(self, db: Session, data: BoletinCreate):
        nuevo = Boletin(
            tipo_documento=data.tipo_documento,
            consecutivo=data.consecutivo,
            fecha=data.fecha or datetime.now(),
            modulo=data.modulo,
            opcion=data.opcion,
            impacto=data.impacto,
            categoria=data.categoria,
            con_documentacion=data.con_documentacion,
            asunto=data.asunto,
            clase_documento=data.clase_documento,
            advertencia=data.advertencia,
            instructivo_descripcion=data.instructivo_descripcion,
            mes=data.mes,
            anio=data.anio,
            archivo=data.archivo,
            fecha_registro=datetime.now()
        )
        return self.repository.create(db, nuevo)
