from datetime import datetime, timezone
from io import BytesIO
from typing import Optional, List
import json
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import desc
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.database import get_db
from app.models.auditoria import AuditLog

BOGOTA_TZ = ZoneInfo("America/Bogota")
UTC_TZ = timezone.utc

router = APIRouter(
    prefix="/auditoria",
    tags=["Auditoría"]
)


class AuditLogResponse(BaseModel):
    oid: int
    fecha_hora: str
    tipo_accion: str
    ip_equipo: str
    nombre_equipo: Optional[str] = None
    usuario_windows_equipo: Optional[str] = None
    modulo: str
    submodulo: str
    usuario: str
    detalle: Optional[str] = None
    payload_json: Optional[dict] = None

    class Config:
        from_attributes = True


def _parse_payload(value: str | None) -> Optional[dict]:
    if not value:
        return None
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {"raw": parsed}
    except Exception:
        return {"raw": value}


def _parse_fecha_inicio(value: str | None) -> Optional[datetime]:
    if not value:
        return None
    normalized = value.replace("Z", "").strip()
    local_dt = datetime.fromisoformat(f"{normalized}T00:00:00" if len(normalized) == 10 else normalized)
    if local_dt.tzinfo is None:
        local_dt = local_dt.replace(tzinfo=BOGOTA_TZ)
    else:
        local_dt = local_dt.astimezone(BOGOTA_TZ)
    return local_dt.astimezone(UTC_TZ).replace(tzinfo=None)


def _parse_fecha_fin(value: str | None) -> Optional[datetime]:
    if not value:
        return None
    normalized = value.replace("Z", "").strip()
    local_dt = datetime.fromisoformat(f"{normalized}T23:59:59.999999" if len(normalized) == 10 else normalized)
    if local_dt.tzinfo is None:
        local_dt = local_dt.replace(tzinfo=BOGOTA_TZ)
    else:
        local_dt = local_dt.astimezone(BOGOTA_TZ)
    return local_dt.astimezone(UTC_TZ).replace(tzinfo=None)


def _to_bogota_string(value: datetime | None) -> str:
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC_TZ)
    return value.astimezone(BOGOTA_TZ).strftime("%Y-%m-%d %H:%M:%S")


@router.get("/", response_model=List[AuditLogResponse])
def listar_auditoria(
    tipo_accion: Optional[str] = Query(None),
    ip_equipo: Optional[str] = Query(None),
    nombre_equipo: Optional[str] = Query(None),
    usuario_windows_equipo: Optional[str] = Query(None),
    modulo: Optional[str] = Query(None),
    submodulo: Optional[str] = Query(None),
    usuario: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(AuditLog)

    if tipo_accion and tipo_accion.strip():
        query = query.filter(AuditLog.tipo_accion.ilike(f"%{tipo_accion.strip()}%"))
    if ip_equipo and ip_equipo.strip():
        query = query.filter(AuditLog.ip_equipo.ilike(f"%{ip_equipo.strip()}%"))
    if nombre_equipo and nombre_equipo.strip():
        query = query.filter(AuditLog.nombre_equipo.ilike(f"%{nombre_equipo.strip()}%"))
    if usuario_windows_equipo and usuario_windows_equipo.strip():
        query = query.filter(AuditLog.usuario_windows_equipo.ilike(f"%{usuario_windows_equipo.strip()}%"))
    if modulo and modulo.strip():
        query = query.filter(AuditLog.modulo.ilike(f"%{modulo.strip()}%"))
    if submodulo and submodulo.strip():
        query = query.filter(AuditLog.submodulo == submodulo.strip())
    if usuario and usuario.strip():
        query = query.filter(AuditLog.usuario.ilike(f"%{usuario.strip()}%"))
    if fecha_inicio:
        try:
            dt_ini = _parse_fecha_inicio(fecha_inicio)
            if dt_ini is not None:
                query = query.filter(AuditLog.fecha_hora >= dt_ini)
        except ValueError:
            pass
    if fecha_fin:
        try:
            dt_fin = _parse_fecha_fin(fecha_fin)
            if dt_fin is not None:
                query = query.filter(AuditLog.fecha_hora <= dt_fin)
        except ValueError:
            pass

    logs = query.order_by(desc(AuditLog.fecha_hora)).limit(1000).all()

    return [
        AuditLogResponse(
            oid=log.oid,
            fecha_hora=_to_bogota_string(log.fecha_hora),
            tipo_accion=log.tipo_accion,
            ip_equipo=log.ip_equipo,
            nombre_equipo=log.nombre_equipo,
            usuario_windows_equipo=log.usuario_windows_equipo,
            modulo=log.modulo,
            submodulo=log.submodulo or "LOGS_SISTEMAS",
            usuario=log.usuario,
            detalle=log.detalle,
            payload_json=_parse_payload(log.payload_json),
        )
        for log in logs
    ]


@router.get("/exportar-excel")
def exportar_auditoria_excel(
    tipo_accion: Optional[str] = Query(None),
    ip_equipo: Optional[str] = Query(None),
    nombre_equipo: Optional[str] = Query(None),
    usuario_windows_equipo: Optional[str] = Query(None),
    modulo: Optional[str] = Query(None),
    submodulo: Optional[str] = Query(None),
    usuario: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(AuditLog)

    if tipo_accion and tipo_accion.strip():
        query = query.filter(AuditLog.tipo_accion.ilike(f"%{tipo_accion.strip()}%"))
    if ip_equipo and ip_equipo.strip():
        query = query.filter(AuditLog.ip_equipo.ilike(f"%{ip_equipo.strip()}%"))
    if nombre_equipo and nombre_equipo.strip():
        query = query.filter(AuditLog.nombre_equipo.ilike(f"%{nombre_equipo.strip()}%"))
    if usuario_windows_equipo and usuario_windows_equipo.strip():
        query = query.filter(AuditLog.usuario_windows_equipo.ilike(f"%{usuario_windows_equipo.strip()}%"))
    if modulo and modulo.strip():
        query = query.filter(AuditLog.modulo.ilike(f"%{modulo.strip()}%"))
    if submodulo and submodulo.strip():
        query = query.filter(AuditLog.submodulo == submodulo.strip())
    if usuario and usuario.strip():
        query = query.filter(AuditLog.usuario.ilike(f"%{usuario.strip()}%"))
    if fecha_inicio:
        try:
            dt_ini = _parse_fecha_inicio(fecha_inicio)
            if dt_ini is not None:
                query = query.filter(AuditLog.fecha_hora >= dt_ini)
        except ValueError:
            pass
    if fecha_fin:
        try:
            dt_fin = _parse_fecha_fin(fecha_fin)
            if dt_fin is not None:
                query = query.filter(AuditLog.fecha_hora <= dt_fin)
        except ValueError:
            pass

    logs = query.order_by(desc(AuditLog.fecha_hora)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoría"

    # Styling
    header_fill = PatternFill(start_color="0778AC", end_color="0778AC", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    headers = [
        "ID Log",
        "Fecha y Hora",
        "Tipo de Acción",
        "Submódulo",
        "Módulo",
        "Usuario App",
        "Usuario Windows",
        "Nombre del Equipo",
        "IP del Equipo",
        "Detalle de la Acción",
    ]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for log in logs:
        ws.append([
            log.oid,
            _to_bogota_string(log.fecha_hora),
            log.tipo_accion,
            log.submodulo or "LOGS_SISTEMAS",
            log.modulo,
            log.usuario,
            log.usuario_windows_equipo or "",
            log.nombre_equipo or "",
            log.ip_equipo,
            log.detalle or ""
        ])

    # Format rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_auditoria_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
